"""Generate editable merchant import workbooks for local demos."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "demo-excel"

HEADERS = {
    "商家信息": ["store_name", "timezone", "address", "phone", "is_open"],
    "服务项目": [
        "service_code",
        "name",
        "price_yuan",
        "duration_minutes",
        "description",
        "is_bookable",
        "is_active",
    ],
    "服务人员": ["staff_code", "name", "gender", "specialty", "is_active"],
    "营业时间": ["weekday", "weekday_name", "open_time", "close_time", "is_closed"],
    "预约规则": ["min_notice_minutes", "cancel_window_minutes", "late_rule"],
    "知识库": ["title", "category", "content", "keywords", "status"],
    "演示客户": ["user_id", "display_name", "phone", "notes"],
    "演示会话": ["conversation_id", "user_id", "subject", "user_message", "assistant_message"],
    "员工排班": ["staff_code", "work_date", "start_time", "end_time", "status", "note"],
    "库存": ["inventory_code", "name", "unit", "quantity", "reorder_level", "unit_cost_yuan", "is_active"],
    "会员": ["member_id", "user_id", "level", "points", "balance_yuan", "joined_at", "status"],
    "优惠券": ["coupon_code", "name", "discount_type", "discount_value", "min_spend_yuan", "valid_from", "valid_to", "status"],
    "支付记录": ["payment_id", "user_id", "appointment_id", "amount_yuan", "payment_method", "status", "paid_at"],
    "演示预约": [
        "appointment_id",
        "user_id",
        "service_code",
        "staff_code",
        "start_time",
        "duration_minutes",
        "status",
    ],
}

DATA: dict[str, dict[str, list[list[Any]]]] = {
    "massage": {
        "商家信息": [["静安舒缓按摩馆", "Asia/Shanghai", "上海市静安区愚园路88号", "021-60001234", 1]],
        "服务项目": [
            ["MASSAGE-001", "肩颈放松", 80, 30, "适合久坐和伏案人群的舒缓项目", 1, 1],
            ["MASSAGE-002", "基础全身放松", 160, 60, "全身基础放松护理", 1, 1],
            ["MASSAGE-003", "足部舒缓", 100, 45, "足部舒缓和日常放松", 1, 1],
            ["MASSAGE-004", "深度放松", 220, 90, "更完整的全身放松体验", 1, 1],
        ],
        "服务人员": [
            ["M-STAFF-001", "林晓", "女", "肩颈放松、精油护理", 1],
            ["M-STAFF-002", "陈安", "男", "足部舒缓、深度放松", 1],
            ["M-STAFF-003", "周宁", "女", "基础全身放松、肩颈放松", 1],
        ],
        "营业时间": [
            [0, "周一", "09:00", "22:00", 0], [1, "周二", "09:00", "22:00", 0],
            [2, "周三", "09:00", "22:00", 0], [3, "周四", "09:00", "22:00", 0],
            [4, "周五", "09:00", "22:00", 0], [5, "周六", "09:00", "23:00", 0],
            [6, "周日", "09:00", "22:00", 0],
        ],
        "预约规则": [[120, 120, "迟到超过15分钟请先联系门店确认；取消或改期请至少提前2小时。"]],
        "知识库": [
            ["营业时间", "门店信息", "门店每天09:00-22:00营业，周六营业至23:00。", "营业时间,开门,几点", "published"],
            ["服务价格", "服务项目", "肩颈放松80元/30分钟，基础全身放松160元/60分钟，足部舒缓100元/45分钟，深度放松220元/90分钟。", "项目,价格,服务", "published"],
            ["预约规则", "预约政策", "预约至少提前2小时提交；取消或改期请至少提前2小时联系门店。", "预约,取消,改期", "published"],
            ["服务说明", "服务项目", "肩颈放松适合久坐、伏案和肩颈紧张人群，具体体验以服务人员现场判断为准。", "肩颈,久坐,放松", "published"],
        ],
        "演示客户": [
            ["massage_customer_001", "张女士", "13800001001", "偏好下午时段，关注肩颈放松"],
            ["massage_customer_002", "李先生", "13800001002", "偏好男服务人员，曾预约深度放松"],
        ],
        "演示会话": [
            ["massage-conv-001", "massage_customer_001", "价格咨询", "肩颈放松多少钱？", "肩颈放松80元，时长30分钟。"],
            ["massage-conv-002", "massage_customer_002", "预约咨询", "明天下午可以做深度放松吗？", "可以查询明天下午的可预约时段，请告诉我大概时间。"],
        ],
        "员工排班": [["M-STAFF-001", "2026-09-01", "09:00", "17:00", "working", "下午重点安排肩颈项目"], ["M-STAFF-002", "2026-09-01", "14:00", "22:00", "working", "晚班"], ["M-STAFF-003", "2026-09-01", "09:00", "18:00", "working", "常规班"]],
        "库存": [["M-INV-001", "按摩精油", "瓶", 24, 8, 68, 1], ["M-INV-002", "一次性床单", "包", 60, 20, 35, 1], ["M-INV-003", "热敷毛巾", "条", 80, 30, 12, 1]],
        "会员": [["M-MEMBER-001", "massage_customer_001", "黄金会员", 680, 320, "2026-01-15", "active"], ["M-MEMBER-002", "massage_customer_002", "普通会员", 120, 80, "2026-04-20", "active"]],
        "优惠券": [["MASSAGE-NEW50", "新客立减券", "fixed", 50, 150, "2026-01-01", "2026-12-31", "active"], ["MASSAGE-OFF10", "会员九折券", "percent", 10, 100, "2026-01-01", "2026-12-31", "active"]],
        "支付记录": [["massage-pay-001", "massage_customer_001", "massage-appt-001", 80, "wechat", "paid", "2026-09-01 14:58"], ["massage-pay-002", "massage_customer_002", "massage-appt-002", 220, "pending", "pending", "2026-09-01 10:00"]],
        "演示预约": [["massage-appt-001", "massage_customer_001", "MASSAGE-001", "M-STAFF-001", "2026-09-01 15:00", 30, "confirmed"], ["massage-appt-002", "massage_customer_002", "MASSAGE-004", "M-STAFF-002", "2026-09-02 19:00", 90, "pending_confirmation"]],
    },
    "beauty": {
        "商家信息": [["静安焕颜美容中心", "Asia/Shanghai", "上海市静安区常德路168号", "021-60005678", 1]],
        "服务项目": [
            ["BEAUTY-001", "基础清洁护理", 128, 60, "日常清洁、补水和基础护理", 1, 1],
            ["BEAUTY-002", "深层补水护理", 198, 75, "适合干燥肌肤的深层补水项目", 1, 1],
            ["BEAUTY-003", "肩颈舒缓护理", 168, 60, "面部护理结合肩颈放松", 1, 1],
            ["BEAUTY-004", "敏感肌修护", 268, 90, "针对敏感状态肌肤的温和修护护理", 1, 1],
        ],
        "服务人员": [
            ["B-STAFF-001", "顾悦", "女", "补水护理、敏感肌修护", 1],
            ["B-STAFF-002", "沈妍", "女", "基础清洁、肩颈舒缓", 1],
            ["B-STAFF-003", "许晴", "女", "敏感肌修护、补水护理", 1],
        ],
        "营业时间": [
            [0, "周一", "10:00", "21:00", 0], [1, "周二", "10:00", "21:00", 0],
            [2, "周三", "10:00", "21:00", 0], [3, "周四", "10:00", "21:00", 0],
            [4, "周五", "10:00", "21:30", 0], [5, "周六", "09:30", "21:30", 0],
            [6, "周日", "09:30", "20:00", 0],
        ],
        "预约规则": [[180, 180, "首次护理建议提前15分钟到店；迟到超过15分钟可能需要调整项目时长。"]],
        "知识库": [
            ["营业时间", "门店信息", "门店周一至周四10:00-21:00营业，周五周六延长营业，周日09:30-20:00营业。", "营业时间,开门,几点", "published"],
            ["服务价格", "服务项目", "基础清洁护理128元/60分钟，深层补水护理198元/75分钟，肩颈舒缓护理168元/60分钟，敏感肌修护268元/90分钟。", "项目,价格,服务", "published"],
            ["护理建议", "服务说明", "首次到店会进行基础肌肤状态沟通；敏感肌修护项目以现场评估和客户确认结果为准。", "敏感肌,修护,评估", "published"],
            ["预约规则", "预约政策", "预约至少提前3小时提交；如需取消或改期，请至少提前3小时联系门店。", "预约,取消,改期", "published"],
        ],
        "演示客户": [
            ["beauty_customer_001", "王女士", "13900002001", "关注补水护理，倾向周末上午"],
            ["beauty_customer_002", "赵女士", "13900002002", "肌肤容易敏感，预约前希望先咨询"],
        ],
        "演示会话": [
            ["beauty-conv-001", "beauty_customer_001", "补水咨询", "干燥肌适合什么项目？", "可以先了解深层补水护理，价格198元，时长75分钟。"],
            ["beauty-conv-002", "beauty_customer_002", "敏感肌咨询", "敏感肌可以做护理吗？", "可以先进行肌肤状态沟通，再由服务人员确认合适的温和护理方案。"],
        ],
        "员工排班": [["B-STAFF-001", "2026-09-05", "09:30", "17:30", "working", "周末早班"], ["B-STAFF-002", "2026-09-05", "11:00", "19:00", "working", "基础护理主班"], ["B-STAFF-003", "2026-09-05", "13:00", "21:30", "working", "敏感肌修护主班"]],
        "库存": [["B-INV-001", "补水面膜", "盒", 36, 10, 88, 1], ["B-INV-002", "舒缓精华", "瓶", 18, 6, 128, 1], ["B-INV-003", "一次性美容床单", "包", 45, 15, 42, 1]],
        "会员": [["B-MEMBER-001", "beauty_customer_001", "白金会员", 920, 560, "2026-02-08", "active"], ["B-MEMBER-002", "beauty_customer_002", "普通会员", 240, 180, "2026-05-11", "active"]],
        "优惠券": [["BEAUTY-NEW80", "新客护理券", "fixed", 80, 198, "2026-01-01", "2026-12-31", "active"], ["BEAUTY-OFF15", "会员护理折扣", "percent", 15, 200, "2026-01-01", "2026-12-31", "active"]],
        "支付记录": [["beauty-pay-001", "beauty_customer_001", "beauty-appt-001", 198, "alipay", "paid", "2026-09-05 10:20"], ["beauty-pay-002", "beauty_customer_002", "beauty-appt-002", 268, "pending", "pending", "2026-09-04 16:00"]],
        "演示预约": [["beauty-appt-001", "beauty_customer_001", "BEAUTY-002", "B-STAFF-001", "2026-09-05 10:30", 75, "confirmed"], ["beauty-appt-002", "beauty_customer_002", "BEAUTY-004", "B-STAFF-003", "2026-09-06 14:00", 90, "pending_confirmation"]],
    },
}


def build_workbook(kind: str, sheets: dict[str, list[list[Any]]]) -> Path:
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_blue = Side(style="thin", color="B7C9D6")

    for sheet_name, headers in HEADERS.items():
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_blue)
        for row in sheets.get(sheet_name, []):
            worksheet.append(row)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 24
        for column in worksheet.columns:
            letter = column[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[letter].width = min(max(max_length + 3, 12), 42)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet_name in {"商家信息", "服务项目", "服务人员", "营业时间", "预约规则", "知识库"}:
            worksheet.sheet_view.showGridLines = False

    instructions = workbook.create_sheet("填写说明", 0)
    instructions.append(["说明", "内容"])
    instructions.append(["用途", "请修改其他工作表中的演示数据后上传到商家后台；不要删除表头。"])
    instructions.append(["金额", "price_yuan 使用元，后台解析后转换为数据库的分（price_cents）。"])
    instructions.append(["布尔值", "is_open、is_bookable、is_active、is_closed 使用 1 或 0。"])
    instructions.append(["星期", "weekday 使用 0=周一至 6=周日。"])
    instructions.append(["时间", "start_time 使用 YYYY-MM-DD HH:MM 格式，例如 2026-09-01 15:00。"])
    instructions.append(["枚举", "知识库 status 建议使用 published；预约 status 可使用 confirmed 或 pending_confirmation。"])
    instructions.append(["关联字段", "演示预约中的 service_code、staff_code 必须分别对应服务项目和服务人员中的编码。"])
    instructions.append(["导入边界", "演示客户、会话、预约用于本地演示；正式导入时后台应对这些数据单独校验。"])
    instructions.freeze_panes = "A2"
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 100
    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in instructions.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for sheet_name in HEADERS:
        worksheet = workbook[sheet_name]
        if sheet_name == "服务项目":
            dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
            worksheet.add_data_validation(dv)
            dv.add(f"F2:G{max(2, worksheet.max_row)}")
        if sheet_name == "服务人员":
            dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
            worksheet.add_data_validation(dv)
            dv.add(f"E2:E{max(2, worksheet.max_row)}")
        if sheet_name == "知识库":
            dv = DataValidation(type="list", formula1='"draft,published,archived"', allow_blank=False)
            worksheet.add_data_validation(dv)
            dv.add(f"E2:E{max(2, worksheet.max_row)}")
        if sheet_name == "演示预约":
            dv = DataValidation(type="list", formula1='"draft,pending_confirmation,confirmed,cancelled"', allow_blank=False)
            worksheet.add_data_validation(dv)
            dv.add(f"G2:G{max(2, worksheet.max_row)}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    target = OUTPUT_DIR / f"{kind}_demo.xlsx"
    workbook.save(target)
    return target


def write_normalized_demo(kind: str, sheets: dict[str, list[list[Any]]]) -> Path:
    """Persist the parsed/normalized demo snapshot consumed by local seed tools."""
    target_dir = ROOT / "demo" / kind
    target_dir.mkdir(parents=True, exist_ok=True)
    for sheet_name, headers in HEADERS.items():
        rows = [dict(zip(headers, row, strict=True)) for row in sheets.get(sheet_name, [])]
        filename = {
            "商家信息": "merchant.json",
            "服务项目": "services.json",
            "服务人员": "staff.json",
            "营业时间": "business_hours.json",
            "预约规则": "appointment_policy.json",
            "知识库": "knowledge.json",
            "演示客户": "customers.json",
            "演示会话": "conversations.json",
            "员工排班": "staff_schedules.json",
            "库存": "inventory.json",
            "会员": "members.json",
            "优惠券": "coupons.json",
            "支付记录": "payments.json",
            "演示预约": "appointments.json",
        }[sheet_name]
        (target_dir / filename).write_text(
            json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    manifest = {
        "demo_kind": kind,
        "source": f"demo-excel/{kind}_demo.xlsx",
        "format": "normalized-json",
        "sheets": {sheet_name: len(sheets.get(sheet_name, [])) for sheet_name in HEADERS},
    }
    (target_dir / "import_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return target_dir


if __name__ == "__main__":
    for demo_kind, demo_sheets in DATA.items():
        print(build_workbook(demo_kind, demo_sheets))
        print(write_normalized_demo(demo_kind, demo_sheets))
